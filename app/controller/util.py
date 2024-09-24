# import csv
import os
from django.conf import settings
# from django.http import HttpResponse
import openpyxl.utils
from ..models import MasterData
import openpyxl
from datetime import datetime

class MakeCSV:

    def __init__(self, input_book):
        self.master_list = MasterData.objects.all()
        self.input_book = input_book
        self.wb = openpyxl.load_workbook(input_book, data_only=True)
        self.ws = self.wb['澤村']
        self.operation_record = ['年月日(作業日)※必須,オーダーNo※必須,オーダー名称,作業フェーズNo※必須,作業フェーズ名称,工数※必須,摘要コード,摘要名,分類１_階層１コード,分類１_階層２コード,分類１_階層１名称,分類１_階層２名称,分類２_階層１コード,分類２_階層２コード,分類２_階層１名称,分類２_階層２名称,備考,勤怠情報ー開始時刻,勤怠情報ー終了時刻,勤怠情報ー休暇時間']
        self.total_work_days_row = 0
        self.search_text_dict = {}
        self.inhouse_work_times = 0

    # '01'の列を検索して、その列番号を返す
    def search_start_date_col(self, keyword):
        result = 7
        for col in self.ws.columns:
            for cell in col:
                try:
                    value = str(cell.value)
                except:
                    continue
                if value == keyword:
                    result = cell.column
        return result

    # '01'の行を検索して、その行番号を返す
    def search_start_date_row(self, keyword):
        result = 2
        for row in self.ws.rows:
            for cell in row:
                try:
                    value = str(cell.value)
                except:
                    continue
                if value == keyword:
                    result = cell.row
        return result

    # 検索文字のある行を取得し、キーにプロジェクト番号、値にプロジェクト名、フェーズ番号、検索文字のある行を持つ辞書を作成
    def get_search_text_dict(self):
        for row in self.ws.iter_rows(min_row=2, max_row=200, min_col=1, max_col=200):
            for col in row:
                for master_data in self.master_list:
                    if col.value == master_data.search_text:
                        project_number = master_data.project_number
                        project_name = master_data.project_name
                        phase_number = master_data.phase_number
                        inhouse_work_flag = master_data.inhouse_work_flag
                        search_text = openpyxl.utils.cell.coordinate_from_string(col.coordinate)[1]
                        self.search_text_dict.update({project_number: [project_name, phase_number, search_text, inhouse_work_flag, 0]})
                    elif col.value == "合計":
                        self.total_work_days_row = col.row

    # search_text_dictの値を更新
    def update_search_text_dict(self, row, col, order_no):
        time_cell = self.ws.cell(row, col).value
        if time_cell is None:
            return
        if not self.search_text_dict[order_no][3]:
            self.search_text_dict[order_no][4] = time_cell
        elif self.search_text_dict[order_no][3] and order_no != self.master_list.filter(project_name='自社作業').first().project_number:
            self.inhouse_work_times -= time_cell
            self.search_text_dict[order_no][4] = time_cell
        else:
            self.search_text_dict[order_no][4] = time_cell

    # CSVレコードを作成
    def create_csv_record(self, work_day):
        work_day_str = work_day.strftime('%Y/%m/%d')
        for key, value in self.search_text_dict.items():
            if value[4] == 0:
                continue
            if value[0] != "自社作業":
                self.operation_record.append(f'{work_day_str},{key},,{value[1]},,{value[4]},,,,,,,,,,,,,,')
            else:
                value[4] += self.inhouse_work_times
                self.operation_record.append(f'{work_day_str},{key},,{value[1]},,{value[4]},,,,,,,,,,,,,,')
                self.inhouse_work_times = 0
        for key, value in self.search_text_dict.items():
            value[4] = 0

# 主な処理
    def operation_record_export(self, year_month_str):
        year_month = datetime.strptime(year_month_str, '%Y-%m').strftime('%Y_%m')
        start_date_col = self.search_start_date_col('01')
        start_date_row = self.search_start_date_row('01')
        self.get_search_text_dict()
        last_col = self.ws.max_column
        for col in range(start_date_col, last_col + 1):
            day_work_sum = self.ws.cell(self.total_work_days_row, col).value
            work_day = self.ws.cell(start_date_row, col).value
            if day_work_sum == 0 or work_day == '計' or work_day is None:
                continue
            # search_text_dictにあるデータ分をCSVに書き込む
            for key, value in self.search_text_dict.items():
                self.update_search_text_dict(value[2], col, key)
            self.create_csv_record(work_day)

        # CSVファイルを作成
        output_dir = os.path.join(settings.BASE_DIR, 'output')
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        file_path = os.path.join(output_dir, 'master_data.csv')
        with open(file_path, 'w') as file:
            for record in self.operation_record:
                file.write("%s\n" % record)
